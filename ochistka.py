import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

INPUT_FILE = "c:\\Users\\harry.boskamp\\OneDrive - Robidus Groep BV\\Digidrat\\відомість.csv"
OUTPUT_FILE = "c:\\Users\\harry.boskamp\\OneDrive - Robidus Groep BV\\Digidrat\\відомість_результат.xlsx"
CHECK_FILE = "c:\\Users\\harry.boskamp\\OneDrive - Robidus Groep BV\\Digidrat\\відомість_переперевірка.xlsx"

# --------------------------------------------------
# Максимальна кількість рядків у Excel-файлі (.xlsx)
MAX_EXCEL_ROWS = 1048576
# 1. Зчитуємо ПЕРШІ 8 РЯДКІВ (заголовок)
# --------------------------------------------------
with open(INPUT_FILE, encoding="cp1251", errors="ignore") as f:
    header_lines = [next(f).rstrip("\n").split(";") for _ in range(8)]


# --------------------------------------------------
# 2. Зчитуємо ОСНОВНІ ДАНІ з 9-го рядка
# --------------------------------------------------
import csv

df = pd.read_csv(
    INPUT_FILE,
    sep=';',
    encoding='cp1251',
    skiprows=8,
    engine='python',
    on_bad_lines='warn',
    dtype=str
)

# Видалення рядка з номерами колонок (перший рядок даних)
df = df.iloc[1:].reset_index(drop=True)



df.columns = df.columns.str.strip()
df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)


# --- ПРАВИЛЬНОЕ ПРЕОБРАЗОВАНИЕ ЧИСЕЛ И ДАТ ---

# Дати
date_columns = [col for col in df.columns if "дата" in col.lower()]
for col in date_columns:
    df[col] = pd.to_datetime(
        df[col],
        errors="coerce",
        format="%d.%m.%Y"
    )

# Система → число
df["Система"] = pd.to_numeric(df["Система"], errors="coerce")

# Конвертація числових колонок (десятковий роздільник — кома)
non_date_cols = [c for c in df.columns if c not in date_columns]
for col in non_date_cols:
    if col == "Система":
        continue
    converted = df[col].str.replace(",", ".", regex=False)
    converted = pd.to_numeric(converted, errors="coerce")
    # Якщо хоча б 50% значень успішно конвертувались — стовпець числовий
    if converted.notna().sum() > df[col].notna().sum() * 0.5:
        df[col] = converted








# --------------------------------------------------
# 4. Заповнення пустих значень у "Система" та "Ідентифікаційний код/номер (ЄДРПОУ)"
# --------------------------------------------------
df["Система"] = df["Система"].ffill()
df["Ідентифікаційний код/номер (ЕДРПОУ)"] = df["Ідентифікаційний код/номер (ЕДРПОУ)"].ffill()

# --------------------------------------------------
# 5. Дебіторська заборгованість → ВИДАЛЯЄМО РЯДОК
# --------------------------------------------------
mask_debit = (
    df["Ознака виду заборгованості"]
    .str.lower()
    .str.contains("дебіторська заборгованість", na=False)
)
df = df.loc[~mask_debit]

# --------------------------------------------------
# 6. Видалення по системам
# --------------------------------------------------

# 2000, 5000, 6000, 7101 + зобов'язання
mask_1 = (
    df["Система"].isin([2000, 5000, 6000, 7101]) &
    (
        df["Ознака виду заборгованості"].isna() |
        df["Ознака виду заборгованості"]
        .str.lower()
        .str.contains("зобов'язання", na=False)
    )
)
df = df.loc[~mask_1]

# 7001, 7003 + кредитна заборгованість
#mask_2 = (
#    df["Система"].isin([7001, 7003]) &
#    df["Ознака виду заборгованості"]
#    .str.lower()
#    .str.contains("кредитна заборгованість", na=False)
#)
#df = df.loc[~mask_2]
# --------------------------------------------------
# 7. ПЕРЕВІРКА
# --------------------------------------------------
systems_to_check = [
    7004, 2000, 5000, 6000,
    7001, 7003, 7101, 7150,
    7160, 9002
]

checks = {
    str(sys): df[df["Система"] == sys]
    for sys in systems_to_check
}

# --------------------------------------------------
# 8. Видалення рядка "ИТОГО"
# --------------------------------------------------
df = df[
    ~df.apply(
        lambda r: r.astype(str)
        .str.contains("ИТОГО", case=False)
        .any(),
        axis=1
    )
]

# --------------------------------------------------
# 9. Запис ОСНОВНОГО ФАЙЛУ з розбиттям
# --------------------------------------------------
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    # запись таблицы: заголовки в строке 9, данные с 10
    df.to_excel(
        writer,
        sheet_name="відомість",
        index=False,
        startrow=8
    )

    # Now get the worksheet after it's been created
    workbook = writer.book
    ws = workbook["відомість"]
    
    # запис шапки (строки 1–8, розбиття по комірках)
    for i, parts in enumerate(header_lines):
        for j, val in enumerate(parts):
            if val:  # пропускаємо порожні комірки
                ws.cell(row=i + 1, column=j + 1).value = val

    # --- ПРАВИЛЬНОЕ ФОРМАТИРОВАНИЕ EXCEL ---
    for idx, col in enumerate(df.columns, start=1):
        # ДАТЫ (по названию колонки)
        if "дата" in col.lower():
            for row_of_cells in ws.iter_rows(
                min_row=10,
                min_col=idx,
                max_col=idx
            ):
                for cell in row_of_cells:
                    if cell.value is not None:
                        cell.number_format = 'dd.mm.yyyy'

    # Handle the second sheet if it exists
    if len(df) > MAX_EXCEL_ROWS:
        ws2 = workbook["Аркуш2"]
        for idx, col in enumerate(df.columns, start=1):
            if "дата" in col.lower():
                for row_of_cells in ws2.iter_rows(
                    min_row=2,
                    min_col=idx,
                    max_col=idx
                ):
                    for cell in row_of_cells:
                        if cell.value is not None:
                            cell.number_format = 'dd.mm.yyyy'
            elif pd.api.types.is_numeric_dtype(df[col]):
                for row_of_cells in ws2.iter_rows(
                    min_row=2,
                    min_col=idx,
                    max_col=idx
                ):
                    for cell in row_of_cells:
                        if cell.value is not None:
                            cell.number_format = numbers.FORMAT_NUMBER





print("✅ Готово")
print("📄 Основний файл:", OUTPUT_FILE)
print("🔍 Файл переперевірки:", CHECK_FILE)